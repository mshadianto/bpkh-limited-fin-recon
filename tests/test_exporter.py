"""Tests for aurix.exporter — ReportExporter."""

import io
import openpyxl
import pytest

from aurix.config import AuditLogEntry
from aurix.exporter import ReportExporter


@pytest.fixture
def audit_log():
    return [
        AuditLogEntry(timestamp="2024-01-01T00:00:00", action="TEST_1", details={"k": "v1"}),
        AuditLogEntry(timestamp="2024-01-01T01:00:00", action="TEST_2", details={"k": "v2"}),
    ]


@pytest.fixture
def excel_buffer(coa_recon_df, txn_detail_df, variance_summary, audit_log):
    return ReportExporter.export_to_excel(
        coa_recon_df, txn_detail_df, variance_summary, audit_log
    )


class TestReportExporter:
    def test_returns_bytesio(self, excel_buffer):
        assert isinstance(excel_buffer, io.BytesIO)

    def test_valid_xlsx(self, excel_buffer):
        wb = openpyxl.load_workbook(excel_buffer)
        assert wb is not None

    def test_has_four_sheets(self, excel_buffer):
        wb = openpyxl.load_workbook(excel_buffer)
        assert len(wb.sheetnames) == 4
        assert "Executive Summary" in wb.sheetnames
        assert "COA Reconciliation" in wb.sheetnames
        assert "Transaction Detail" in wb.sheetnames
        assert "Audit Trail" in wb.sheetnames

    def test_coa_sheet_has_data_rows(self, excel_buffer, coa_recon_df):
        wb = openpyxl.load_workbook(excel_buffer)
        ws = wb["COA Reconciliation"]
        # Header row + data rows; at minimum more than 1 row
        assert ws.max_row > 1

    def test_audit_trail_has_entries(self, excel_buffer, audit_log):
        wb = openpyxl.load_workbook(excel_buffer)
        ws = wb["Audit Trail"]
        # At least header + audit entries
        assert ws.max_row >= len(audit_log) + 1

    def test_empty_data(self, variance_summary):
        """Export with zero-row DataFrames (correct schema) should not crash."""
        import pandas as pd
        empty_coa = pd.DataFrame(columns=[
            "COA", "Account_Name", "Manual_Debit", "Manual_Credit", "Manual_Net",
            "Daftra_Debit", "Daftra_Credit", "Daftra_Net",
            "Debit_Variance", "Credit_Variance", "Net_Variance", "Status",
        ])
        empty_txn = pd.DataFrame(columns=[
            "Date", "COA", "Account_Name", "Description",
            "Debit", "Credit", "Net", "Reference", "RowID", "Source",
        ])
        result = ReportExporter.export_to_excel(empty_coa, empty_txn, variance_summary, [])
        assert isinstance(result, io.BytesIO)
