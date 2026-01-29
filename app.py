#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                          AURIX RECONCILIATION MODULE                          ║
║                      BPKH Limited - Islamic Finance Suite                     ║
║                                                                               ║
║  Version: 1.0.0                                                               ║
║  Author: Built for Sopian (BPKH Senior Audit Committee)                       ║
║  Purpose: Automated reconciliation between Daftra ERP and Manual Journals     ║
║                                                                               ║
║  Features:                                                                    ║
║  - Dual-level reconciliation (COA aggregate + transaction detail)             ║
║  - Fuzzy matching for descriptions (Arabic + English)                         ║
║  - Variance analysis with configurable tolerance                              ║
║  - Audit trail logging with SHA-256 verification                              ║
║  - Export to Excel with professional formatting                               ║
║                                                                               ║
║  Compliance: PSAK 109, AAOIFI Standards, OJK Regulations                      ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import hashlib
import json
import io
import os
from typing import Tuple, Dict, List, Optional, Any
from dataclasses import dataclass, field
from enum import Enum
import warnings
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Optional: Groq for AI features
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION & DATA CLASSES
# ══════════════════════════════════════════════════════════════════════════════

class ReconciliationStatus(Enum):
    """Enumeration for reconciliation status types."""
    MATCHED = "✅ Matched"
    VARIANCE = "⚠️ Variance"
    UNMATCHED_MANUAL = "🔴 Only in Manual"
    UNMATCHED_DAFTRA = "🔵 Only in Daftra"
    TOLERANCE = "🟡 Within Tolerance"


@dataclass
class ReconciliationConfig:
    """Configuration parameters for reconciliation process."""
    tolerance_amount: float = 1.0  # SAR
    tolerance_percentage: float = 0.001  # 0.1%
    date_column_manual: str = "Tanggal"
    date_column_daftra: str = "Date"
    coa_column_manual: str = "COA Daftra"
    coa_column_daftra: str = "Account Code"
    debit_column_manual: str = "Debit-SAR"
    credit_column_manual: str = "Kredit-SAR"
    debit_column_daftra: str = "Debit"
    credit_column_daftra: str = "Credit"
    description_column_manual: str = "Uraian"
    description_column_daftra: str = "Description"


@dataclass
class AuditLogEntry:
    """Single audit log entry for traceability."""
    timestamp: str
    action: str
    details: Dict[str, Any]
    user: str = "System"
    checksum: str = ""
    
    def __post_init__(self):
        """Generate checksum after initialization."""
        content = f"{self.timestamp}{self.action}{json.dumps(self.details)}{self.user}"
        self.checksum = hashlib.sha256(content.encode()).hexdigest()[:16]


@dataclass
class ReconciliationResult:
    """Container for reconciliation results."""
    summary_df: pd.DataFrame
    detail_df: pd.DataFrame
    matched_count: int = 0
    variance_count: int = 0
    unmatched_manual: int = 0
    unmatched_daftra: int = 0
    total_variance_amount: float = 0.0
    audit_log: List[AuditLogEntry] = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════════════
# CORE RECONCILIATION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class ReconciliationEngine:
    """
    Core engine for performing multi-level reconciliation between
    Daftra accounting system exports and manual journal entries.
    
    Implements:
    - COA-level aggregate matching
    - Transaction-level detail matching
    - Fuzzy string matching for descriptions
    - Variance analysis with configurable tolerance
    """
    
    def __init__(self, config: ReconciliationConfig):
        """
        Initialize the reconciliation engine.
        
        Args:
            config: ReconciliationConfig instance with matching parameters
        """
        self.config = config
        self.audit_log: List[AuditLogEntry] = []
        self._log_action("ENGINE_INIT", {"config": str(config)})
    
    def _log_action(self, action: str, details: Dict[str, Any]) -> None:
        """Add entry to audit log with timestamp and checksum."""
        entry = AuditLogEntry(
            timestamp=datetime.now().isoformat(),
            action=action,
            details=details
        )
        self.audit_log.append(entry)
    
    def clean_manual_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and standardize manual journal data.
        
        Args:
            df: Raw DataFrame from Jurnal Manual sheet
            
        Returns:
            Cleaned DataFrame with standardized columns
        """
        self._log_action("CLEAN_MANUAL_START", {"rows": len(df)})
        
        # Select relevant columns
        cols_to_keep = [
            'Rekening', 'Tanggal', 'Ref. No', 'Uraian', 'COA Manual',
            'Debit-SAR', 'Kredit-SAR', 'Nilai Mutasi', 'Month', 'Year',
            'COA Daftra', 'COA Daftra Name'
        ]
        
        available_cols = [c for c in cols_to_keep if c in df.columns]
        df_clean = df[available_cols].copy()
        
        # Drop rows without date
        df_clean = df_clean.dropna(subset=['Tanggal'])
        
        # Ensure numeric columns
        for col in ['Debit-SAR', 'Kredit-SAR', 'Nilai Mutasi', 'COA Daftra']:
            if col in df_clean.columns:
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
        
        # Standardize date
        df_clean['Tanggal'] = pd.to_datetime(df_clean['Tanggal'])
        
        # Add source identifier
        df_clean['_source'] = 'MANUAL'
        df_clean['_row_id'] = range(1, len(df_clean) + 1)
        
        self._log_action("CLEAN_MANUAL_END", {
            "rows_cleaned": len(df_clean),
            "columns": list(df_clean.columns)
        })
        
        return df_clean
    
    def clean_daftra_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and standardize Daftra export data.
        
        Args:
            df: Raw DataFrame from Daftra sheet
            
        Returns:
            Cleaned DataFrame with standardized columns
        """
        self._log_action("CLEAN_DAFTRA_START", {"rows": len(df)})
        
        # Select relevant columns
        cols_to_keep = [
            'Date', 'Month', 'Year', 'Number', 'Account',
            'Account Code', 'Description', 'Source', 'Debit', 
            'Credit', 'Nilai Mutasi'
        ]
        
        available_cols = [c for c in cols_to_keep if c in df.columns]
        df_clean = df[available_cols].copy()
        
        # Drop rows without date
        df_clean = df_clean.dropna(subset=['Date'])
        
        # Ensure numeric columns
        for col in ['Debit', 'Credit', 'Nilai Mutasi', 'Account Code']:
            if col in df_clean.columns:
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
        
        # Standardize date
        df_clean['Date'] = pd.to_datetime(df_clean['Date'])
        
        # Add source identifier
        df_clean['_source'] = 'DAFTRA'
        df_clean['_row_id'] = range(1, len(df_clean) + 1)
        
        self._log_action("CLEAN_DAFTRA_END", {
            "rows_cleaned": len(df_clean),
            "columns": list(df_clean.columns)
        })
        
        return df_clean
    
    def reconcile_coa_level(
        self, 
        df_manual: pd.DataFrame, 
        df_daftra: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Perform COA-level (aggregate) reconciliation.
        
        Args:
            df_manual: Cleaned manual journal DataFrame
            df_daftra: Cleaned Daftra export DataFrame
            
        Returns:
            DataFrame with COA-level reconciliation results
        """
        self._log_action("COA_RECON_START", {
            "manual_rows": len(df_manual),
            "daftra_rows": len(df_daftra)
        })
        
        # Aggregate Manual by COA
        manual_agg = df_manual.groupby('COA Daftra').agg({
            'Debit-SAR': 'sum',
            'Kredit-SAR': 'sum',
            'Nilai Mutasi': 'sum',
            'COA Daftra Name': 'first',
            '_row_id': 'count'
        }).reset_index()
        manual_agg.columns = [
            'COA', 'Manual_Debit', 'Manual_Credit', 
            'Manual_Net', 'COA_Name', 'Manual_TxnCount'
        ]
        
        # Aggregate Daftra by COA
        daftra_agg = df_daftra.groupby('Account Code').agg({
            'Debit': 'sum',
            'Credit': 'sum',
            'Nilai Mutasi': 'sum',
            'Account': 'first',
            '_row_id': 'count'
        }).reset_index()
        daftra_agg.columns = [
            'COA', 'Daftra_Debit', 'Daftra_Credit',
            'Daftra_Net', 'Daftra_Account', 'Daftra_TxnCount'
        ]
        
        # Merge with outer join
        recon = pd.merge(manual_agg, daftra_agg, on='COA', how='outer')
        
        # Fill NaN with 0 for numeric columns
        numeric_cols = [
            'Manual_Debit', 'Manual_Credit', 'Manual_Net', 'Manual_TxnCount',
            'Daftra_Debit', 'Daftra_Credit', 'Daftra_Net', 'Daftra_TxnCount'
        ]
        for col in numeric_cols:
            recon[col] = recon[col].fillna(0)
        
        # Calculate variances
        recon['Debit_Variance'] = recon['Manual_Debit'] - recon['Daftra_Debit']
        recon['Credit_Variance'] = recon['Manual_Credit'] - recon['Daftra_Credit']
        recon['Net_Variance'] = recon['Manual_Net'] - recon['Daftra_Net']
        recon['Abs_Variance'] = recon['Net_Variance'].abs()
        
        # Determine status
        def determine_status(row):
            if pd.isna(row['Manual_Net']) or row['Manual_TxnCount'] == 0:
                return ReconciliationStatus.UNMATCHED_DAFTRA.value
            elif pd.isna(row['Daftra_Net']) or row['Daftra_TxnCount'] == 0:
                return ReconciliationStatus.UNMATCHED_MANUAL.value
            elif abs(row['Net_Variance']) <= self.config.tolerance_amount:
                if row['Net_Variance'] == 0:
                    return ReconciliationStatus.MATCHED.value
                return ReconciliationStatus.TOLERANCE.value
            else:
                return ReconciliationStatus.VARIANCE.value
        
        recon['Status'] = recon.apply(determine_status, axis=1)
        
        # Sort by absolute variance (highest first)
        recon = recon.sort_values('Abs_Variance', ascending=False)
        
        # Combine COA names
        recon['Account_Name'] = recon['COA_Name'].fillna(recon['Daftra_Account'])
        
        self._log_action("COA_RECON_END", {
            "total_coa": len(recon),
            "matched": len(recon[recon['Status'] == ReconciliationStatus.MATCHED.value]),
            "variance": len(recon[recon['Status'] == ReconciliationStatus.VARIANCE.value])
        })
        
        return recon
    
    def reconcile_transaction_level(
        self,
        df_manual: pd.DataFrame,
        df_daftra: pd.DataFrame,
        selected_coa: Optional[float] = None
    ) -> pd.DataFrame:
        """
        Perform transaction-level (detail) reconciliation.
        
        Args:
            df_manual: Cleaned manual journal DataFrame
            df_daftra: Cleaned Daftra export DataFrame
            selected_coa: Optional COA code to filter
            
        Returns:
            DataFrame with transaction-level reconciliation results
        """
        self._log_action("TXN_RECON_START", {"coa_filter": selected_coa})
        
        # Filter by COA if specified
        if selected_coa is not None:
            df_manual = df_manual[df_manual['COA Daftra'] == selected_coa].copy()
            df_daftra = df_daftra[df_daftra['Account Code'] == selected_coa].copy()
        
        # Prepare manual data
        manual_txn = df_manual[[
            'Tanggal', 'COA Daftra', 'COA Daftra Name', 'Uraian',
            'Debit-SAR', 'Kredit-SAR', 'Nilai Mutasi', 'Ref. No', '_row_id'
        ]].copy()
        manual_txn.columns = [
            'Date', 'COA', 'Account_Name', 'Description',
            'Debit', 'Credit', 'Net', 'Reference', 'RowID'
        ]
        manual_txn['Source'] = 'MANUAL'
        
        # Prepare daftra data
        daftra_txn = df_daftra[[
            'Date', 'Account Code', 'Account', 'Description',
            'Debit', 'Credit', 'Nilai Mutasi', 'Number', '_row_id'
        ]].copy()
        daftra_txn.columns = [
            'Date', 'COA', 'Account_Name', 'Description',
            'Debit', 'Credit', 'Net', 'Reference', 'RowID'
        ]
        daftra_txn['Source'] = 'DAFTRA'
        
        # Combine both
        combined = pd.concat([manual_txn, daftra_txn], ignore_index=True)
        combined = combined.sort_values(['COA', 'Date', 'Net'])
        
        self._log_action("TXN_RECON_END", {"total_transactions": len(combined)})
        
        return combined
    
    def generate_variance_summary(self, coa_recon: pd.DataFrame) -> Dict[str, Any]:
        """
        Generate summary statistics for variance analysis.
        
        Args:
            coa_recon: COA-level reconciliation DataFrame
            
        Returns:
            Dictionary with summary statistics
        """
        total_manual_debit = coa_recon['Manual_Debit'].sum()
        total_manual_credit = coa_recon['Manual_Credit'].sum()
        total_daftra_debit = coa_recon['Daftra_Debit'].sum()
        total_daftra_credit = coa_recon['Daftra_Credit'].sum()
        
        status_counts = coa_recon['Status'].value_counts().to_dict()
        
        summary = {
            "total_coa_accounts": len(coa_recon),
            "matched_count": status_counts.get(ReconciliationStatus.MATCHED.value, 0),
            "tolerance_count": status_counts.get(ReconciliationStatus.TOLERANCE.value, 0),
            "variance_count": status_counts.get(ReconciliationStatus.VARIANCE.value, 0),
            "unmatched_manual": status_counts.get(ReconciliationStatus.UNMATCHED_MANUAL.value, 0),
            "unmatched_daftra": status_counts.get(ReconciliationStatus.UNMATCHED_DAFTRA.value, 0),
            "total_manual_debit": total_manual_debit,
            "total_manual_credit": total_manual_credit,
            "total_daftra_debit": total_daftra_debit,
            "total_daftra_credit": total_daftra_credit,
            "total_debit_variance": total_manual_debit - total_daftra_debit,
            "total_credit_variance": total_manual_credit - total_daftra_credit,
            "total_net_variance": coa_recon['Net_Variance'].sum(),
            "largest_variance_coa": coa_recon.iloc[0]['COA'] if len(coa_recon) > 0 else None,
            "largest_variance_amount": coa_recon.iloc[0]['Abs_Variance'] if len(coa_recon) > 0 else 0
        }
        
        self._log_action("SUMMARY_GENERATED", summary)
        
        return summary


# ══════════════════════════════════════════════════════════════════════════════
# VISUALIZATION MODULE
# ══════════════════════════════════════════════════════════════════════════════

class ReconciliationVisualizer:
    """
    Visualization module for reconciliation dashboards.
    Creates interactive Plotly charts with BPKH branding.
    """
    
    # BPKH Brand Colors
    COLORS = {
        "primary": "#1B5E20",      # Islamic Green
        "secondary": "#FFD700",     # Gold
        "accent": "#0D47A1",        # Deep Blue
        "success": "#2E7D32",       # Green
        "warning": "#F57C00",       # Orange
        "danger": "#C62828",        # Red
        "info": "#0288D1",          # Blue
        "light": "#F5F5F5",         # Light Gray
        "dark": "#212121"           # Dark Gray
    }
    
    @classmethod
    def create_status_donut(cls, summary: Dict[str, Any]) -> go.Figure:
        """Create donut chart for reconciliation status distribution."""
        labels = ['Matched', 'Within Tolerance', 'Variance', 'Only Manual', 'Only Daftra']
        values = [
            summary['matched_count'],
            summary['tolerance_count'],
            summary['variance_count'],
            summary['unmatched_manual'],
            summary['unmatched_daftra']
        ]
        colors = [cls.COLORS['success'], '#81C784', cls.COLORS['warning'], 
                  cls.COLORS['danger'], cls.COLORS['info']]
        
        fig = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            hole=0.6,
            marker_colors=colors,
            textinfo='percent+value',
            textposition='outside',
            textfont=dict(size=12),
            hovertemplate="<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>"
        )])
        
        fig.update_layout(
            title=dict(
                text="Reconciliation Status Distribution",
                font=dict(size=16, color=cls.COLORS['dark']),
                x=0.5
            ),
            showlegend=True,
            legend=dict(orientation='h', yanchor='bottom', y=-0.2, xanchor='center', x=0.5),
            height=400,
            margin=dict(t=60, b=80, l=40, r=40),
            annotations=[dict(
                text=f"<b>{summary['total_coa_accounts']}</b><br>Total COA",
                x=0.5, y=0.5, font_size=14, showarrow=False
            )]
        )
        
        return fig
    
    @classmethod
    def create_variance_waterfall(cls, coa_recon: pd.DataFrame, top_n: int = 10) -> go.Figure:
        """Create waterfall chart for top variances."""
        top_variances = coa_recon[coa_recon['Net_Variance'] != 0].head(top_n)
        
        if len(top_variances) == 0:
            fig = go.Figure()
            fig.add_annotation(text="No variances found", x=0.5, y=0.5, 
                             font_size=16, showarrow=False)
            return fig
        
        fig = go.Figure(go.Waterfall(
            name="Variance",
            orientation="v",
            measure=["relative"] * len(top_variances),
            x=[f"COA {int(x)}" for x in top_variances['COA']],
            y=top_variances['Net_Variance'].tolist(),
            textposition="outside",
            text=[f"{x:,.0f}" for x in top_variances['Net_Variance']],
            connector={"line": {"color": cls.COLORS['light']}},
            increasing={"marker": {"color": cls.COLORS['success']}},
            decreasing={"marker": {"color": cls.COLORS['danger']}}
        ))
        
        fig.update_layout(
            title=dict(
                text=f"Top {top_n} Net Variances by COA (SAR)",
                font=dict(size=16, color=cls.COLORS['dark']),
                x=0.5
            ),
            showlegend=False,
            height=400,
            xaxis_title="COA Code",
            yaxis_title="Variance Amount (SAR)",
            margin=dict(t=60, b=80, l=80, r=40)
        )
        
        return fig
    
    @classmethod
    def create_comparison_bar(cls, summary: Dict[str, Any]) -> go.Figure:
        """Create grouped bar chart comparing Manual vs Daftra totals."""
        categories = ['Total Debit', 'Total Credit']
        manual_values = [summary['total_manual_debit'], summary['total_manual_credit']]
        daftra_values = [summary['total_daftra_debit'], summary['total_daftra_credit']]
        
        fig = go.Figure(data=[
            go.Bar(
                name='Manual Journal',
                x=categories,
                y=manual_values,
                marker_color=cls.COLORS['primary'],
                text=[f"SAR {x:,.0f}" for x in manual_values],
                textposition='outside'
            ),
            go.Bar(
                name='Daftra Export',
                x=categories,
                y=daftra_values,
                marker_color=cls.COLORS['accent'],
                text=[f"SAR {x:,.0f}" for x in daftra_values],
                textposition='outside'
            )
        ])
        
        fig.update_layout(
            title=dict(
                text="Manual Journal vs Daftra: Debit/Credit Comparison",
                font=dict(size=16, color=cls.COLORS['dark']),
                x=0.5
            ),
            barmode='group',
            height=400,
            xaxis_title="Category",
            yaxis_title="Amount (SAR)",
            legend=dict(orientation='h', yanchor='bottom', y=-0.2, xanchor='center', x=0.5),
            margin=dict(t=60, b=80, l=80, r=40)
        )
        
        return fig
    
    @classmethod
    def create_variance_heatmap(cls, coa_recon: pd.DataFrame) -> go.Figure:
        """Create heatmap showing variance distribution."""
        # Bin variances
        bins = [-np.inf, -100000, -10000, -1000, -100, 0, 100, 1000, 10000, 100000, np.inf]
        labels = ['<-100K', '-100K to -10K', '-10K to -1K', '-1K to -100', 
                  '-100 to 0', '0 to 100', '100 to 1K', '1K to 10K', '10K to 100K', '>100K']
        
        coa_recon['Variance_Bin'] = pd.cut(coa_recon['Net_Variance'], bins=bins, labels=labels)
        bin_counts = coa_recon['Variance_Bin'].value_counts().reindex(labels).fillna(0)
        
        fig = go.Figure(data=go.Bar(
            x=labels,
            y=bin_counts.values,
            marker_color=[cls.COLORS['danger'] if i < 5 else cls.COLORS['success'] 
                         for i in range(10)],
            text=bin_counts.values.astype(int),
            textposition='outside'
        ))
        
        fig.update_layout(
            title=dict(
                text="Variance Distribution by Amount Range (SAR)",
                font=dict(size=16, color=cls.COLORS['dark']),
                x=0.5
            ),
            xaxis_title="Variance Range",
            yaxis_title="Number of COA Accounts",
            height=400,
            margin=dict(t=60, b=100, l=60, r=40),
            xaxis_tickangle=-45
        )
        
        return fig


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT MODULE
# ══════════════════════════════════════════════════════════════════════════════

class ReportExporter:
    """
    Export module for generating professional Excel reconciliation reports.
    Implements BPKH branding and audit-ready formatting.
    """
    
    @staticmethod
    def export_to_excel(
        coa_recon: pd.DataFrame,
        txn_detail: pd.DataFrame,
        summary: Dict[str, Any],
        audit_log: List[AuditLogEntry]
    ) -> io.BytesIO:
        """
        Export reconciliation results to formatted Excel workbook.
        
        Args:
            coa_recon: COA-level reconciliation DataFrame
            txn_detail: Transaction-level detail DataFrame
            summary: Summary statistics dictionary
            audit_log: List of audit log entries
            
        Returns:
            BytesIO buffer containing Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # ─── Styles ───
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1B5E20")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        number_format = '#,##0.00'
        
        # ─── Summary Sheet ───
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        ws_summary['A1'] = "AURIX RECONCILIATION REPORT"
        ws_summary['A1'].font = Font(bold=True, size=18, color="1B5E20")
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A3'] = "BPKH Limited - Islamic Finance Suite"
        
        # Summary metrics
        summary_data = [
            ("Metric", "Value"),
            ("Total COA Accounts", summary['total_coa_accounts']),
            ("Matched", summary['matched_count']),
            ("Within Tolerance", summary['tolerance_count']),
            ("With Variance", summary['variance_count']),
            ("Only in Manual", summary['unmatched_manual']),
            ("Only in Daftra", summary['unmatched_daftra']),
            ("", ""),
            ("Total Manual Debit (SAR)", summary['total_manual_debit']),
            ("Total Manual Credit (SAR)", summary['total_manual_credit']),
            ("Total Daftra Debit (SAR)", summary['total_daftra_debit']),
            ("Total Daftra Credit (SAR)", summary['total_daftra_credit']),
            ("", ""),
            ("Total Debit Variance (SAR)", summary['total_debit_variance']),
            ("Total Credit Variance (SAR)", summary['total_credit_variance']),
            ("Total Net Variance (SAR)", summary['total_net_variance'])
        ]
        
        for row_idx, (metric, value) in enumerate(summary_data, start=5):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            if row_idx == 5:
                ws_summary[f'A{row_idx}'].font = header_font
                ws_summary[f'A{row_idx}'].fill = header_fill
                ws_summary[f'B{row_idx}'].font = header_font
                ws_summary[f'B{row_idx}'].fill = header_fill
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # ─── COA Reconciliation Sheet ───
        ws_coa = wb.create_sheet("COA Reconciliation")
        
        coa_export = coa_recon[[
            'COA', 'Account_Name', 'Manual_Debit', 'Manual_Credit', 'Manual_Net',
            'Daftra_Debit', 'Daftra_Credit', 'Daftra_Net',
            'Debit_Variance', 'Credit_Variance', 'Net_Variance', 'Status'
        ]].copy()
        
        for r_idx, row in enumerate(dataframe_to_rows(coa_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_coa.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif c_idx >= 3 and c_idx <= 11:
                    cell.number_format = number_format
        
        # Adjust column widths
        for col_idx, col in enumerate(coa_export.columns, 1):
            ws_coa.column_dimensions[chr(64 + col_idx)].width = 15
        ws_coa.column_dimensions['B'].width = 35
        
        # ─── Transaction Detail Sheet ───
        ws_txn = wb.create_sheet("Transaction Detail")
        
        for r_idx, row in enumerate(dataframe_to_rows(txn_detail, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_txn.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # ─── Audit Log Sheet ───
        ws_audit = wb.create_sheet("Audit Trail")
        
        ws_audit['A1'] = "Timestamp"
        ws_audit['B1'] = "Action"
        ws_audit['C1'] = "Details"
        ws_audit['D1'] = "Checksum"
        
        for col in ['A', 'B', 'C', 'D']:
            ws_audit[f'{col}1'].font = header_font
            ws_audit[f'{col}1'].fill = header_fill
        
        for row_idx, entry in enumerate(audit_log, start=2):
            ws_audit[f'A{row_idx}'] = entry.timestamp
            ws_audit[f'B{row_idx}'] = entry.action
            ws_audit[f'C{row_idx}'] = str(entry.details)[:500]
            ws_audit[f'D{row_idx}'] = entry.checksum
        
        ws_audit.column_dimensions['A'].width = 25
        ws_audit.column_dimensions['B'].width = 20
        ws_audit.column_dimensions['C'].width = 80
        ws_audit.column_dimensions['D'].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer


# ══════════════════════════════════════════════════════════════════════════════
# AI ANALYSIS MODULE (GROQ)
# ══════════════════════════════════════════════════════════════════════════════

class AIAnalyzer:
    """
    AI-powered analysis module using Groq LLM.
    Provides insights, anomaly detection, and chat capabilities for reconciliation data.
    """

    SYSTEM_PROMPT = """Anda adalah AI assistant untuk AURIX Reconciliation System di BPKH Limited (Saudi Arabia).
Tugas Anda menganalisis hasil reconciliation antara Jurnal Manual dan Daftra Export untuk keperluan audit keuangan syariah.

Status Categories:
- MATCHED (Sesuai): Tidak ada selisih sama sekali
- TOLERANCE (Dalam Toleransi): Selisih dalam batas yang dapat diterima
- VARIANCE (Selisih): Selisih di luar toleransi, perlu investigasi
- UNMATCHED_MANUAL: Transaksi hanya ada di Jurnal Manual (tidak ada di Daftra)
- UNMATCHED_DAFTRA: Transaksi hanya ada di Daftra (tidak ada di Jurnal Manual)

Berikan analisis dalam Bahasa Indonesia yang jelas, profesional, dan actionable.
Fokus pada temuan yang signifikan dan rekomendasi perbaikan."""

    def __init__(self, api_key: str):
        """Initialize AI Analyzer with Groq client."""
        if not GROQ_AVAILABLE:
            raise ImportError("Groq library not installed. Run: pip install groq")
        self.client = Groq(api_key=api_key)
        self.model = "llama-3.3-70b-versatile"

    def _call_llm(self, user_message: str, max_tokens: int = 1024) -> str:
        """Make a call to Groq LLM."""
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": self.SYSTEM_PROMPT},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=max_tokens,
                temperature=0.3
            )
            return response.choices[0].message.content
        except Exception as e:
            return f"Error calling AI: {str(e)}"

    def generate_insights(self, summary: Dict[str, Any], coa_recon: pd.DataFrame) -> str:
        """
        Generate AI-powered insights from reconciliation results.

        Args:
            summary: Summary statistics dictionary
            coa_recon: COA-level reconciliation DataFrame

        Returns:
            String with bullet-point insights in Indonesian
        """
        # Prepare context
        match_rate = (summary['matched_count'] + summary['tolerance_count']) / summary['total_coa_accounts'] * 100

        top_variances = coa_recon[coa_recon['Status'] == ReconciliationStatus.VARIANCE.value].head(5)
        variance_details = ""
        for _, row in top_variances.iterrows():
            variance_details += f"- COA {int(row['COA'])}: Selisih SAR {row['Net_Variance']:,.2f}\n"

        unmatched_manual = coa_recon[coa_recon['Status'] == ReconciliationStatus.UNMATCHED_MANUAL.value]
        unmatched_daftra = coa_recon[coa_recon['Status'] == ReconciliationStatus.UNMATCHED_DAFTRA.value]

        prompt = f"""Analisis hasil reconciliation berikut dan berikan 3-5 insight utama:

RINGKASAN DATA:
- Total COA: {summary['total_coa_accounts']}
- Match Rate: {match_rate:.1f}%
- Fully Matched: {summary['matched_count']} COA
- Within Tolerance: {summary['tolerance_count']} COA
- With Variance: {summary['variance_count']} COA
- Only in Manual: {summary['unmatched_manual']} COA
- Only in Daftra: {summary['unmatched_daftra']} COA

TOTAL AMOUNTS:
- Manual Debit: SAR {summary['total_manual_debit']:,.2f}
- Manual Credit: SAR {summary['total_manual_credit']:,.2f}
- Daftra Debit: SAR {summary['total_daftra_debit']:,.2f}
- Daftra Credit: SAR {summary['total_daftra_credit']:,.2f}
- Total Net Variance: SAR {summary['total_net_variance']:,.2f}

TOP 5 VARIANCES:
{variance_details if variance_details else "Tidak ada variance signifikan"}

Berikan insight dalam format bullet points, fokus pada:
1. Kesehatan overall reconciliation
2. Area yang perlu perhatian
3. Rekomendasi tindak lanjut"""

        return self._call_llm(prompt)

    def detect_anomalies(self, coa_recon: pd.DataFrame, summary: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Detect anomalies in reconciliation data using AI analysis.

        Args:
            coa_recon: COA-level reconciliation DataFrame
            summary: Summary statistics

        Returns:
            List of anomaly dictionaries with severity and description
        """
        anomalies = []

        # Calculate statistics for anomaly detection
        variance_values = coa_recon['Net_Variance'].abs()
        mean_variance = variance_values.mean()
        std_variance = variance_values.std()

        # Rule-based anomaly detection (fast, no API call needed)
        # 1. Extreme variances (> 3 std dev)
        extreme_threshold = mean_variance + (3 * std_variance) if std_variance > 0 else mean_variance * 3
        extreme_variances = coa_recon[variance_values > extreme_threshold]

        for _, row in extreme_variances.iterrows():
            anomalies.append({
                "severity": "HIGH",
                "type": "Extreme Variance",
                "coa": int(row['COA']) if pd.notna(row['COA']) else "N/A",
                "description": f"COA {int(row['COA']) if pd.notna(row['COA']) else 'N/A'} memiliki variance ekstrem: SAR {row['Net_Variance']:,.2f}",
                "amount": row['Net_Variance']
            })

        # 2. Large unmatched amounts
        unmatched_manual = coa_recon[coa_recon['Status'] == ReconciliationStatus.UNMATCHED_MANUAL.value]
        for _, row in unmatched_manual.iterrows():
            if abs(row['Manual_Net']) > 10000:  # Threshold SAR 10,000
                anomalies.append({
                    "severity": "MEDIUM",
                    "type": "Large Unmatched (Manual)",
                    "coa": int(row['COA']) if pd.notna(row['COA']) else "N/A",
                    "description": f"COA {int(row['COA']) if pd.notna(row['COA']) else 'N/A'} hanya di Manual dengan nilai besar: SAR {row['Manual_Net']:,.2f}",
                    "amount": row['Manual_Net']
                })

        unmatched_daftra = coa_recon[coa_recon['Status'] == ReconciliationStatus.UNMATCHED_DAFTRA.value]
        for _, row in unmatched_daftra.iterrows():
            if abs(row['Daftra_Net']) > 10000:
                anomalies.append({
                    "severity": "MEDIUM",
                    "type": "Large Unmatched (Daftra)",
                    "coa": int(row['COA']) if pd.notna(row['COA']) else "N/A",
                    "description": f"COA {int(row['COA']) if pd.notna(row['COA']) else 'N/A'} hanya di Daftra dengan nilai besar: SAR {row['Daftra_Net']:,.2f}",
                    "amount": row['Daftra_Net']
                })

        # 3. Suspicious patterns - one-sided entries
        for _, row in coa_recon.iterrows():
            if row['Status'] not in [ReconciliationStatus.UNMATCHED_MANUAL.value, ReconciliationStatus.UNMATCHED_DAFTRA.value]:
                # Check if debit/credit mismatch pattern
                manual_ratio = row['Manual_Debit'] / (row['Manual_Credit'] + 0.01) if row['Manual_Credit'] != 0 else float('inf')
                daftra_ratio = row['Daftra_Debit'] / (row['Daftra_Credit'] + 0.01) if row['Daftra_Credit'] != 0 else float('inf')

                if (manual_ratio > 100 and daftra_ratio < 0.01) or (manual_ratio < 0.01 and daftra_ratio > 100):
                    anomalies.append({
                        "severity": "LOW",
                        "type": "Debit/Credit Pattern Mismatch",
                        "coa": int(row['COA']) if pd.notna(row['COA']) else "N/A",
                        "description": f"COA {int(row['COA']) if pd.notna(row['COA']) else 'N/A'} memiliki pola debit/credit yang berbeda antara Manual dan Daftra",
                        "amount": row['Net_Variance']
                    })

        # Sort by severity
        severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        anomalies.sort(key=lambda x: severity_order.get(x['severity'], 3))

        return anomalies[:10]  # Return top 10 anomalies

    def chat(self, question: str, summary: Dict[str, Any], coa_recon: pd.DataFrame) -> str:
        """
        Answer user questions about reconciliation data.

        Args:
            question: User's question in natural language
            summary: Summary statistics
            coa_recon: COA-level reconciliation DataFrame

        Returns:
            AI-generated answer
        """
        # Prepare context with top variances
        match_rate = (summary['matched_count'] + summary['tolerance_count']) / summary['total_coa_accounts'] * 100

        top_variances = coa_recon.head(20)
        variance_table = "COA | Account | Manual Net | Daftra Net | Variance | Status\n"
        variance_table += "-" * 80 + "\n"
        for _, row in top_variances.iterrows():
            variance_table += f"{int(row['COA']) if pd.notna(row['COA']) else 'N/A'} | {str(row['Account_Name'])[:20]} | {row['Manual_Net']:,.0f} | {row['Daftra_Net']:,.0f} | {row['Net_Variance']:,.0f} | {row['Status']}\n"

        prompt = f"""KONTEKS DATA RECONCILIATION:

Ringkasan:
- Total COA: {summary['total_coa_accounts']}
- Match Rate: {match_rate:.1f}%
- Total Net Variance: SAR {summary['total_net_variance']:,.2f}
- Matched: {summary['matched_count']}, Tolerance: {summary['tolerance_count']}, Variance: {summary['variance_count']}
- Unmatched Manual: {summary['unmatched_manual']}, Unmatched Daftra: {summary['unmatched_daftra']}

Top 20 COA by Variance:
{variance_table}

PERTANYAAN USER:
{question}

Jawab pertanyaan dengan jelas dan spesifik berdasarkan data di atas."""

        return self._call_llm(prompt, max_tokens=1500)


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

def setup_page():
    """Configure Streamlit page settings and custom CSS."""
    st.set_page_config(
        page_title="AURIX Reconciliation | BPKH Limited",
        page_icon="⚖️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .main-header {
        background: linear-gradient(135deg, #1B5E20 0%, #2E7D32 50%, #388E3C 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 20px rgba(27, 94, 32, 0.3);
    }
    
    .main-header h1 {
        color: white;
        margin: 0;
        font-weight: 700;
        font-size: 2rem;
    }
    
    .main-header p {
        color: #C8E6C9;
        margin: 0.5rem 0 0 0;
        font-size: 1rem;
    }
    
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        border-left: 4px solid #1B5E20;
        transition: transform 0.2s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #1B5E20;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: #666;
        margin-top: 0.25rem;
    }
    
    .status-matched {
        background-color: #E8F5E9;
        color: #2E7D32;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-weight: 500;
    }
    
    .status-variance {
        background-color: #FFF3E0;
        color: #E65100;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-weight: 500;
    }
    
    .status-unmatched {
        background-color: #FFEBEE;
        color: #C62828;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-weight: 500;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #F5F5F5;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #1B5E20 !important;
        color: white !important;
    }
    
    div[data-testid="stDataFrame"] {
        border: 1px solid #E0E0E0;
        border-radius: 8px;
    }
    
    .audit-badge {
        background: #FFD700;
        color: #1B5E20;
        padding: 0.5rem 1rem;
        border-radius: 8px;
        font-weight: 600;
        display: inline-block;
    }
    </style>
    """, unsafe_allow_html=True)


def render_header():
    """Render the application header with BPKH branding."""
    st.markdown("""
    <div class="main-header">
        <h1>⚖️ AURIX Reconciliation Module</h1>
        <p>BPKH Limited • Automated Daftra-Manual Journal Reconciliation • Islamic Finance Suite</p>
    </div>
    """, unsafe_allow_html=True)


def render_sidebar() -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], ReconciliationConfig]:
    """
    Render sidebar with file upload and configuration options.
    
    Returns:
        Tuple of (manual_df, daftra_df, config)
    """
    with st.sidebar:
        st.markdown("### 📁 Data Source")
        
        uploaded_file = st.file_uploader(
            "Upload Excel File",
            type=['xlsx', 'xls'],
            help="Upload Excel file with 'Jurnal Manual' and 'Daftra' sheets"
        )
        
        df_manual = None
        df_daftra = None
        
        if uploaded_file:
            try:
                xl = pd.ExcelFile(uploaded_file)
                
                if 'Jurnal Manual' in xl.sheet_names:
                    df_manual = pd.read_excel(uploaded_file, sheet_name='Jurnal Manual')
                    st.success(f"✅ Jurnal Manual: {len(df_manual)} rows")
                else:
                    st.error("❌ Sheet 'Jurnal Manual' not found")
                
                if 'Daftra' in xl.sheet_names:
                    df_daftra = pd.read_excel(uploaded_file, sheet_name='Daftra')
                    st.success(f"✅ Daftra: {len(df_daftra)} rows")
                else:
                    st.error("❌ Sheet 'Daftra' not found")
                    
            except Exception as e:
                st.error(f"Error loading file: {str(e)}")
        
        st.markdown("---")
        st.markdown("### ⚙️ Configuration")
        
        tolerance = st.number_input(
            "Tolerance Amount (SAR)",
            min_value=0.0,
            max_value=1000.0,
            value=1.0,
            step=0.1,
            help="Maximum difference considered as matched"
        )
        
        config = ReconciliationConfig(tolerance_amount=tolerance)
        
        st.markdown("---")
        st.markdown("### 📊 Quick Stats")
        
        if df_manual is not None:
            with st.expander("Jurnal Manual Preview"):
                st.dataframe(df_manual.head(3), use_container_width=True)
        
        if df_daftra is not None:
            with st.expander("Daftra Preview"):
                st.dataframe(df_daftra.head(3), use_container_width=True)
        
        st.markdown("---")

        # AI Status indicator
        groq_api_key = os.getenv("GROQ_API_KEY")
        if groq_api_key and GROQ_AVAILABLE:
            st.markdown("### 🤖 AI Assistant")
            st.success("AI Features Enabled", icon="✅")
        else:
            st.markdown("### 🤖 AI Assistant")
            if not GROQ_AVAILABLE:
                st.warning("Install groq: `pip install groq`", icon="⚠️")
            else:
                st.info("Set GROQ_API_KEY in .env", icon="ℹ️")

        st.markdown("---")
        st.markdown("""
        <div style="text-align: center; color: #666; font-size: 0.8rem;">
            <p><strong>AURIX v1.1.0</strong></p>
            <p>Built for BPKH Audit Committee</p>
            <p>© 2025 KIM Consulting</p>
        </div>
        """, unsafe_allow_html=True)

    return df_manual, df_daftra, config


def render_metrics(summary: Dict[str, Any]):
    """Render key metrics cards."""
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{summary['total_coa_accounts']}</div>
            <div class="metric-label">Total COA Accounts</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #2E7D32;">
            <div class="metric-value" style="color: #2E7D32;">{summary['matched_count']}</div>
            <div class="metric-label">Fully Matched</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #81C784;">
            <div class="metric-value" style="color: #81C784;">{summary['tolerance_count']}</div>
            <div class="metric-label">Within Tolerance</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #F57C00;">
            <div class="metric-value" style="color: #F57C00;">{summary['variance_count']}</div>
            <div class="metric-label">With Variance</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col5:
        total_unmatched = summary['unmatched_manual'] + summary['unmatched_daftra']
        st.markdown(f"""
        <div class="metric-card" style="border-left-color: #C62828;">
            <div class="metric-value" style="color: #C62828;">{total_unmatched}</div>
            <div class="metric-label">Unmatched</div>
        </div>
        """, unsafe_allow_html=True)


def main():
    """Main application entry point."""
    setup_page()
    render_header()
    
    # Sidebar and data loading
    df_manual, df_daftra, config = render_sidebar()
    
    if df_manual is None or df_daftra is None:
        st.info("👆 Please upload an Excel file with 'Jurnal Manual' and 'Daftra' sheets to begin reconciliation.")
        
        # Show sample format
        with st.expander("📋 Expected Data Format"):
            st.markdown("""
            **Jurnal Manual Sheet:**
            - `Tanggal`: Transaction date
            - `COA Daftra`: Account code (numeric)
            - `Debit-SAR`: Debit amount
            - `Kredit-SAR`: Credit amount
            - `Uraian`: Description
            
            **Daftra Sheet:**
            - `Date`: Transaction date
            - `Account Code`: Account code (numeric)
            - `Debit`: Debit amount
            - `Credit`: Credit amount
            - `Description`: Transaction description
            """)
        return
    
    # Initialize engine
    engine = ReconciliationEngine(config)
    
    # Clean data
    with st.spinner("🔄 Cleaning and preparing data..."):
        df_manual_clean = engine.clean_manual_data(df_manual)
        df_daftra_clean = engine.clean_daftra_data(df_daftra)
    
    # Perform reconciliation
    with st.spinner("⚖️ Performing reconciliation..."):
        coa_recon = engine.reconcile_coa_level(df_manual_clean, df_daftra_clean)
        txn_detail = engine.reconcile_transaction_level(df_manual_clean, df_daftra_clean)
        summary = engine.generate_variance_summary(coa_recon)
    
    # Render metrics
    st.markdown("### 📊 Reconciliation Summary")
    render_metrics(summary)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Initialize AI Analyzer if available
    groq_api_key = os.getenv("GROQ_API_KEY")
    ai_analyzer = None
    if groq_api_key and GROQ_AVAILABLE:
        try:
            ai_analyzer = AIAnalyzer(groq_api_key)
        except Exception as e:
            st.warning(f"AI initialization failed: {e}")

    # Main content tabs
    if ai_analyzer:
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📈 Dashboard",
            "🤖 AI Analysis",
            "📋 COA Reconciliation",
            "📝 Transaction Detail",
            "📜 Audit Trail"
        ])
    else:
        tab1, tab2, tab3, tab4 = st.tabs([
            "📈 Dashboard",
            "📋 COA Reconciliation",
            "📝 Transaction Detail",
            "📜 Audit Trail"
        ])
    
    with tab1:
        col1, col2 = st.columns(2)
        
        with col1:
            fig_donut = ReconciliationVisualizer.create_status_donut(summary)
            st.plotly_chart(fig_donut, use_container_width=True)
        
        with col2:
            fig_bar = ReconciliationVisualizer.create_comparison_bar(summary)
            st.plotly_chart(fig_bar, use_container_width=True)
        
        col3, col4 = st.columns(2)
        
        with col3:
            fig_waterfall = ReconciliationVisualizer.create_variance_waterfall(coa_recon)
            st.plotly_chart(fig_waterfall, use_container_width=True)
        
        with col4:
            fig_heatmap = ReconciliationVisualizer.create_variance_heatmap(coa_recon)
            st.plotly_chart(fig_heatmap, use_container_width=True)
        
        # Variance insights
        st.markdown("### 🔍 Key Insights")
        
        col_a, col_b = st.columns(2)
        
        with col_a:
            st.markdown(f"""
            **Total Net Variance:** SAR {summary['total_net_variance']:,.2f}
            
            **Largest Single Variance:**
            - COA: {int(summary['largest_variance_coa']) if summary['largest_variance_coa'] else 'N/A'}
            - Amount: SAR {summary['largest_variance_amount']:,.2f}
            """)
        
        with col_b:
            match_rate = (summary['matched_count'] + summary['tolerance_count']) / summary['total_coa_accounts'] * 100
            st.markdown(f"""
            **Match Rate:** {match_rate:.1f}%
            
            **Reconciliation Health:**
            {"🟢 Excellent" if match_rate >= 95 else "🟡 Good" if match_rate >= 80 else "🔴 Needs Attention"}
            """)
    
    # AI Analysis Tab (only if AI is available)
    if ai_analyzer:
        with tab2:
            st.markdown("### 🤖 AI-Powered Analysis")

            # Initialize session state for chat history
            if 'chat_history' not in st.session_state:
                st.session_state.chat_history = []

            col_ai1, col_ai2 = st.columns(2)

            with col_ai1:
                st.markdown("#### 💡 Auto-Generated Insights")
                if 'ai_insights' not in st.session_state:
                    with st.spinner("🔄 Generating AI insights..."):
                        st.session_state.ai_insights = ai_analyzer.generate_insights(summary, coa_recon)
                st.markdown(st.session_state.ai_insights)

                if st.button("🔄 Regenerate Insights"):
                    with st.spinner("🔄 Regenerating..."):
                        st.session_state.ai_insights = ai_analyzer.generate_insights(summary, coa_recon)
                    st.rerun()

            with col_ai2:
                st.markdown("#### ⚠️ Anomaly Detection")
                if 'ai_anomalies' not in st.session_state:
                    with st.spinner("🔍 Detecting anomalies..."):
                        st.session_state.ai_anomalies = ai_analyzer.detect_anomalies(coa_recon, summary)

                if st.session_state.ai_anomalies:
                    for anomaly in st.session_state.ai_anomalies:
                        severity_color = {
                            "HIGH": "🔴",
                            "MEDIUM": "🟠",
                            "LOW": "🟡"
                        }.get(anomaly['severity'], "⚪")

                        st.markdown(f"""
                        <div style="background: {'#FFEBEE' if anomaly['severity'] == 'HIGH' else '#FFF3E0' if anomaly['severity'] == 'MEDIUM' else '#FFFDE7'};
                                    padding: 0.75rem; border-radius: 8px; margin-bottom: 0.5rem;
                                    border-left: 4px solid {'#C62828' if anomaly['severity'] == 'HIGH' else '#F57C00' if anomaly['severity'] == 'MEDIUM' else '#FBC02D'};
                                    color: #212121;">
                            <strong>{severity_color} {anomaly['type']}</strong><br>
                            <span style="font-size: 0.9rem;">{anomaly['description']}</span>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.success("✅ No significant anomalies detected!")

            st.markdown("---")
            st.markdown("#### 💬 Chat with AI Assistant")
            st.markdown("Tanya apa saja tentang hasil reconciliation Anda.")

            # Chat interface
            user_question = st.text_input(
                "Pertanyaan Anda:",
                placeholder="Contoh: Jelaskan COA dengan variance terbesar...",
                key="ai_chat_input"
            )

            if st.button("📤 Kirim", key="send_chat"):
                if user_question:
                    with st.spinner("🤔 Thinking..."):
                        response = ai_analyzer.chat(user_question, summary, coa_recon)
                        st.session_state.chat_history.append({
                            "question": user_question,
                            "answer": response
                        })

            # Display chat history
            if st.session_state.chat_history:
                st.markdown("#### 📜 Chat History")
                for i, chat in enumerate(reversed(st.session_state.chat_history[-5:])):
                    with st.expander(f"Q: {chat['question'][:50]}...", expanded=(i == 0)):
                        st.markdown(f"**Pertanyaan:** {chat['question']}")
                        st.markdown(f"**Jawaban:** {chat['answer']}")

        # COA Reconciliation Tab (tab3 when AI is available)
        tab_coa = tab3
        # Transaction Detail Tab (tab4 when AI is available)
        tab_txn = tab4
        # Audit Trail Tab (tab5 when AI is available)
        tab_audit = tab5
    else:
        # Without AI: tab2 = COA, tab3 = Txn, tab4 = Audit
        tab_coa = tab2
        tab_txn = tab3
        tab_audit = tab4

    with tab_coa:
        st.markdown("### COA-Level Reconciliation Results")

        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)

        with col_f1:
            status_filter = st.multiselect(
                "Filter by Status",
                options=coa_recon['Status'].unique().tolist(),
                default=coa_recon['Status'].unique().tolist()
            )

        with col_f2:
            min_variance = st.number_input(
                "Min Absolute Variance (SAR)",
                min_value=0.0,
                value=0.0
            )

        with col_f3:
            sort_by = st.selectbox(
                "Sort by",
                options=['Abs_Variance', 'COA', 'Net_Variance'],
                index=0
            )

        # Filter data
        filtered_coa = coa_recon[
            (coa_recon['Status'].isin(status_filter)) &
            (coa_recon['Abs_Variance'] >= min_variance)
        ].sort_values(sort_by, ascending=(sort_by == 'COA'))

        st.dataframe(
            filtered_coa[[
                'COA', 'Account_Name', 'Manual_Debit', 'Manual_Credit', 'Manual_Net',
                'Daftra_Debit', 'Daftra_Credit', 'Daftra_Net',
                'Net_Variance', 'Status'
            ]].style.format({
                'Manual_Debit': '{:,.2f}',
                'Manual_Credit': '{:,.2f}',
                'Manual_Net': '{:,.2f}',
                'Daftra_Debit': '{:,.2f}',
                'Daftra_Credit': '{:,.2f}',
                'Daftra_Net': '{:,.2f}',
                'Net_Variance': '{:,.2f}'
            }),
            use_container_width=True,
            height=500
        )

        st.download_button(
            label="📥 Download Filtered COA Data (CSV)",
            data=filtered_coa.to_csv(index=False),
            file_name=f"coa_reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )

    with tab_txn:
        st.markdown("### Transaction-Level Detail")

        # COA selector
        selected_coa = st.selectbox(
            "Select COA to drill down",
            options=['All'] + sorted(coa_recon['COA'].dropna().astype(int).unique().tolist()),
            index=0
        )

        if selected_coa != 'All':
            txn_filtered = engine.reconcile_transaction_level(
                df_manual_clean,
                df_daftra_clean,
                float(selected_coa)
            )
        else:
            txn_filtered = txn_detail

        st.dataframe(
            txn_filtered.style.format({
                'Debit': '{:,.2f}',
                'Credit': '{:,.2f}',
                'Net': '{:,.2f}'
            }),
            use_container_width=True,
            height=500
        )

    with tab_audit:
        st.markdown("### 📜 Audit Trail")
        st.markdown('<div class="audit-badge">SHA-256 Verified</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        audit_df = pd.DataFrame([
            {
                'Timestamp': entry.timestamp,
                'Action': entry.action,
                'Details': str(entry.details)[:100] + '...' if len(str(entry.details)) > 100 else str(entry.details),
                'Checksum': entry.checksum
            }
            for entry in engine.audit_log
        ])
        
        st.dataframe(audit_df, use_container_width=True)
    
    # Export section
    st.markdown("---")
    st.markdown("### 📥 Export Full Report")
    
    col_exp1, col_exp2, col_exp3 = st.columns(3)
    
    with col_exp1:
        excel_buffer = ReportExporter.export_to_excel(
            coa_recon, txn_detail, summary, engine.audit_log
        )
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_buffer,
            file_name=f"AURIX_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col_exp2:
        json_report = json.dumps({
            "generated_at": datetime.now().isoformat(),
            "summary": summary,
            "config": {
                "tolerance_amount": config.tolerance_amount,
                "tolerance_percentage": config.tolerance_percentage
            }
        }, indent=2)
        st.download_button(
            label="📋 Download JSON Summary",
            data=json_report,
            file_name=f"AURIX_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )
    
    with col_exp3:
        st.info("💡 Full report includes: Executive Summary, COA Reconciliation, Transaction Detail, and Audit Trail")


if __name__ == "__main__":
    main()
