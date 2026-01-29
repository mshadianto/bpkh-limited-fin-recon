"""Export module for AURIX Reconciliation reports."""

import io
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any

from aurix.config import AuditLogEntry


class ReportExporter:
    """Export module for generating professional Excel reconciliation reports."""

    @staticmethod
    def export_to_excel(
        coa_recon: pd.DataFrame,
        txn_detail: pd.DataFrame,
        summary: Dict[str, Any],
        audit_log: List[AuditLogEntry]
    ) -> io.BytesIO:
        """Export reconciliation results to formatted Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1B5E20")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        number_format = '#,##0.00'

        # Executive Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary['A1'] = "AURIX RECONCILIATION REPORT"
        ws_summary['A1'].font = Font(bold=True, size=18, color="1B5E20")
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A3'] = "BPKH Limited - Islamic Finance Suite"

        summary_data = [
            ("Metric", "Value"),
            ("Total COA Accounts", summary['total_coa_accounts']),
            ("Matched", summary['matched_count']),
            ("Within Tolerance", summary['tolerance_count']),
            ("With Variance", summary['variance_count']),
            ("Only in Manual", summary['unmatched_manual']),
            ("Only in Daftra", summary['unmatched_daftra']),
            ("", ""),
            ("Total Manual Debit (SAR)", summary['total_manual_debit']),
            ("Total Manual Credit (SAR)", summary['total_manual_credit']),
            ("Total Daftra Debit (SAR)", summary['total_daftra_debit']),
            ("Total Daftra Credit (SAR)", summary['total_daftra_credit']),
            ("", ""),
            ("Total Debit Variance (SAR)", summary['total_debit_variance']),
            ("Total Credit Variance (SAR)", summary['total_credit_variance']),
            ("Total Net Variance (SAR)", summary['total_net_variance'])
        ]

        for row_idx, (metric, value) in enumerate(summary_data, start=5):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            if row_idx == 5:
                ws_summary[f'A{row_idx}'].font = header_font
                ws_summary[f'A{row_idx}'].fill = header_fill
                ws_summary[f'B{row_idx}'].font = header_font
                ws_summary[f'B{row_idx}'].fill = header_fill

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # COA Reconciliation Sheet
        ws_coa = wb.create_sheet("COA Reconciliation")
        coa_export = coa_recon[[
            'COA', 'Account_Name', 'Manual_Debit', 'Manual_Credit', 'Manual_Net',
            'Daftra_Debit', 'Daftra_Credit', 'Daftra_Net',
            'Debit_Variance', 'Credit_Variance', 'Net_Variance', 'Status'
        ]].copy()

        for r_idx, row in enumerate(dataframe_to_rows(coa_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_coa.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                elif c_idx >= 3 and c_idx <= 11:
                    cell.number_format = number_format

        for col_idx, col in enumerate(coa_export.columns, 1):
            ws_coa.column_dimensions[chr(64 + col_idx)].width = 15
        ws_coa.column_dimensions['B'].width = 35

        # Transaction Detail Sheet
        ws_txn = wb.create_sheet("Transaction Detail")
        for r_idx, row in enumerate(dataframe_to_rows(txn_detail, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_txn.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # Audit Log Sheet
        ws_audit = wb.create_sheet("Audit Trail")
        ws_audit['A1'] = "Timestamp"
        ws_audit['B1'] = "Action"
        ws_audit['C1'] = "Details"
        ws_audit['D1'] = "Checksum"

        for col in ['A', 'B', 'C', 'D']:
            ws_audit[f'{col}1'].font = header_font
            ws_audit[f'{col}1'].fill = header_fill

        for row_idx, entry in enumerate(audit_log, start=2):
            ws_audit[f'A{row_idx}'] = entry.timestamp
            ws_audit[f'B{row_idx}'] = entry.action
            ws_audit[f'C{row_idx}'] = str(entry.details)[:500]
            ws_audit[f'D{row_idx}'] = entry.checksum

        ws_audit.column_dimensions['A'].width = 25
        ws_audit.column_dimensions['B'].width = 20
        ws_audit.column_dimensions['C'].width = 80
        ws_audit.column_dimensions['D'].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
